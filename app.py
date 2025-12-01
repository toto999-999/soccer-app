import streamlit as st
import pandas as pd
import numpy as np
import re
import warnings
import openpyxl
from collections import defaultdict, deque
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.ensemble import RandomForestRegressor
from scipy.stats import poisson
from xgboost import XGBClassifier
import tempfile
import os

# 경고 무시
warnings.filterwarnings('ignore')

# 페이지 기본 설정 (제목, 레이아웃)
st.set_page_config(page_title="AI 축구 분석기 (Web)", layout="wide")

# ==========================================
# [1] 데이터 로드 클래스 (수정됨)
# ==========================================
class DataLoader:
    def __init__(self):
        self.handi_map = {} 
        self.color_map = {} 

    def load_excel_data(self, uploaded_file):
        try:
            # Streamlit에서 업로드된 파일을 임시 파일로 저장 (openpyxl 색상 인식을 위해)
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(uploaded_file.getvalue())
                tmp_path = tmp.name

            # 1. 색상 정보 로드
            self._load_colors_with_openpyxl(tmp_path)
            # 2. 핸디캡 데이터 로드
            self._load_handicap_data(tmp_path)

            # 3. 데이터 값 로드
            raw_df = pd.read_excel(tmp_path, sheet_name='배당변경(일반)', header=0)
            
            # 임시 파일 삭제
            os.remove(tmp_path)
            
            processed_rows = []
            n_rows = len(raw_df)
            
            for i in range(0, n_rows, 2):
                if i + 1 >= n_rows: break
                
                row1 = raw_df.iloc[i]
                row2 = raw_df.iloc[i+1] 
                
                try:
                    w_odd = pd.to_numeric(row1.iloc[7], errors='coerce')
                    if pd.isna(w_odd): continue 

                    year = row1.iloc[0]
                    time_str = row1.iloc[1]
                    league = str(row1.iloc[2]).strip()
                    home = str(row1.iloc[3]).strip()
                    away = str(row1.iloc[5]).strip()
                    result_wdl = str(row1.iloc[10]).strip()
                    
                    score_str = str(row2.iloc[10]).strip()
                    h_score, a_score = np.nan, np.nan
                    
                    if ':' in score_str:
                        try:
                            parts = score_str.split(':')
                            h_score = int(parts[0])
                            a_score = int(parts[1])
                        except: pass
                    
                    is_scheduled = result_wdl not in ['승', '무', '패']
                    
                    processed_rows.append({
                        '년도': year,
                        '경기시간': time_str,
                        'League': league,
                        'Home': home,
                        'Away': away,
                        '승': w_odd,
                        '무': pd.to_numeric(row1.iloc[8], errors='coerce'),
                        '패': pd.to_numeric(row1.iloc[9], errors='coerce'),
                        '결과_승무패': result_wdl if not is_scheduled else None,
                        '홈팀득점': h_score,
                        '원정팀득점': a_score,
                        '경기유형': '예정' if is_scheduled else '과거'
                    })
                    
                except Exception:
                    continue

            if not processed_rows:
                return None

            df = pd.DataFrame(processed_rows)
            
            try:
                date_strs = df['년도'].astype(str) + ' ' + df['경기시간'].astype(str)
                date_strs_clean = date_strs.apply(lambda x: re.sub(r'\([가-힣]\)', '', x))
                df['datetime'] = pd.to_datetime(date_strs_clean, errors='coerce')
            except:
                df['datetime'] = pd.NaT
            
            df = df.sort_values(by='datetime').reset_index(drop=True)
            return df
            
        except Exception as e:
            st.error(f"데이터 로드 중 에러 발생: {e}")
            return None

    def _load_handicap_data(self, file_path):
        try:
            df_h = pd.read_excel(file_path, sheet_name='일반핸디', header=0)
            self.handi_map = {}
            for idx, row in df_h.iterrows():
                try:
                    if len(row) < 18: continue
                    h_team = str(row.iloc[3]).strip()
                    a_team = str(row.iloc[5]).strip()
                    
                    key = f"{h_team}vs{a_team}"
                    self.handi_map[key] = {
                        'handi_val': row.iloc[14], 
                        'h_odd': pd.to_numeric(row.iloc[15], errors='coerce'),
                        'd_odd': pd.to_numeric(row.iloc[16], errors='coerce'),
                        'l_odd': pd.to_numeric(row.iloc[17], errors='coerce')
                    }
                except: continue
        except: self.handi_map = {}

    def _load_colors_with_openpyxl(self, file_path):
        self.color_map = {}
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if '일반핸디' not in wb.sheetnames: return
            ws = wb['일반핸디']
            
            for row in ws.iter_rows(min_row=2):
                try:
                    h_team = str(row[3].value).strip()
                    a_team = str(row[5].value).strip()
                    if not h_team or h_team == 'None': continue
                    
                    key = f"{h_team}vs{a_team}"
                    
                    def is_colored(cell):
                        if cell.fill and cell.fill.start_color:
                            c_code = str(cell.fill.start_color.index)
                            rgb = str(cell.fill.start_color.rgb)
                            if c_code in ['00000000', 'FFFFFFFF', '0', 'None']: return False
                            if rgb in ['00000000', 'FFFFFFFF']: return False
                            return True
                        return False

                    self.color_map[key] = {
                        'N_W': is_colored(row[7]), 'N_D': is_colored(row[8]), 'N_L': is_colored(row[9]),
                        'H_W': is_colored(row[15]), 'H_D': is_colored(row[16]), 'H_L': is_colored(row[17])
                    }
                except Exception: continue
            wb.close()
        except Exception: pass

    def check_trap(self, home, away, win_odd):
        key = f"{home}vs{away}"
        info = self.handi_map.get(key)
        if not info: return ""
        try:
            h_odd = info['h_odd']
            if win_odd < 1.6 and h_odd > 2.5: return "🚨함정(핸디괴리)"
            if win_odd < 1.8 and h_odd < 2.0: return "⚠️핸디특이"
        except: pass
        return ""

# ==========================================
# [2] 피처 엔지니어링 & 모델링 함수
# ==========================================
def engineer_features(df):
    if df is None: return None
    home_stats = defaultdict(lambda: {'goals': deque(maxlen=5), 'conceded': deque(maxlen=5), 'points': deque(maxlen=5)})
    away_stats = defaultdict(lambda: {'goals': deque(maxlen=5), 'conceded': deque(maxlen=5), 'points': deque(maxlen=5)})
    home_only_stats = defaultdict(lambda: deque(maxlen=5)) 
    away_only_stats = defaultdict(lambda: deque(maxlen=5))
    h2h_stats = defaultdict(lambda: deque(maxlen=3))
    elo_ratings = defaultdict(lambda: 1500)
    last_match_date = {} 
    k_factor = 40 
    new_features = []
    
    for index, row in df.iterrows():
        home, away = row['Home'], row['Away']
        current_date = row['datetime']
        h_rest = 7; a_rest = 7
        if pd.notnull(current_date):
            if home in last_match_date: h_rest = (current_date - last_match_date[home]).days
            if away in last_match_date: a_rest = (current_date - last_match_date[away]).days
        
        home_elo, away_elo = elo_ratings[home], elo_ratings[away]
        prob_elo = 1 / (1 + 10 ** (-(home_elo - away_elo) / 400))
        def get_avg(dq): return sum(dq)/len(dq) if dq else 0
        
        features = {
            'elo_diff': home_elo - away_elo,
            'prob_elo': prob_elo,
            'home_form': get_avg(home_stats[home]['points']),
            'away_form': get_avg(away_stats[away]['points']),
            'home_venue_perf': get_avg(home_only_stats[home]),
            'away_venue_perf': get_avg(away_only_stats[away]),
            'home_rest': min(h_rest, 30),
            'away_rest': min(a_rest, 30),
            'att_def_diff': (get_avg(home_stats[home]['goals']) + get_avg(away_stats[away]['conceded'])) - 
                            (get_avg(away_stats[away]['goals']) + get_avg(home_stats[home]['conceded'])),
            'h2h_balance': sum(h2h_stats[f"{home}_vs_{away}"]) if h2h_stats[f"{home}_vs_{away}"] else 0,
            'implied_prob_w': (1 / row['승']) if row['승'] > 1 else 0,
            'odds_gap': row['패'] - row['승'] if (row['승'] > 0 and row['패'] > 0) else 0
        }
        new_features.append(features)
        
        if row['경기유형'] == '과거':
            res = row['결과_승무패']
            hg = row['홈팀득점'] if pd.notna(row['홈팀득점']) else 0
            ag = row['원정팀득점'] if pd.notna(row['원정팀득점']) else 0
            if pd.notnull(current_date):
                last_match_date[home] = current_date; last_match_date[away] = current_date
            h_pt, a_pt = 1, 1; actual = 0.5; h2h_pt = 0
            if res == '승': h_pt, a_pt = 3, 0; actual = 1.0; h2h_pt = 1
            elif res == '패': h_pt, a_pt = 0, 3; actual = 0.0; h2h_pt = -1
            change = k_factor * (actual - prob_elo)
            elo_ratings[home] += change; elo_ratings[away] -= change
            home_stats[home]['points'].append(h_pt); home_stats[home]['goals'].append(hg); home_stats[home]['conceded'].append(ag)
            away_stats[away]['points'].append(a_pt); away_stats[away]['goals'].append(ag); away_stats[away]['conceded'].append(hg)
            home_only_stats[home].append(h_pt); away_only_stats[away].append(a_pt)
            h2h_stats[f"{home}_vs_{away}"].append(h2h_pt); h2h_stats[f"{away}_vs_{home}"].append(-h2h_pt)
    return pd.concat([df, pd.DataFrame(new_features)], axis=1)

def train_models_func(df):
    if df is None or df.empty: return None
    train = df[df['경기유형'] == '과거'].copy()
    features = ['elo_diff', 'prob_elo', 'home_form', 'away_form', 'home_venue_perf', 'away_venue_perf', 
                'home_rest', 'away_rest',
                'att_def_diff', 'h2h_balance', 'implied_prob_w', 'odds_gap']
    X = train[features].fillna(0)
    y = train['결과_승무패']
    if len(y) < 10: return None
    
    le = LabelEncoder()
    try: y_enc = le.fit_transform(y)
    except: return None
    
    scaler = StandardScaler()
    X_scaled = pd.DataFrame(scaler.fit_transform(X), columns=features, index=X.index)
    
    # [수정] AI 모델 경량화 (서버 과부하 방지)
    # n_estimators: 500 -> 50 (나무 개수를 1/10로 줄임)
    # max_depth: 5 -> 3 (생각의 깊이를 얕게 함)
    model = XGBClassifier(n_estimators=50, learning_rate=0.05, max_depth=3, 
                          use_label_encoder=False, eval_metric='mlogloss', n_jobs=1, random_state=42)
    model.fit(X_scaled, y_enc)
    
    hg_model, ag_model = None, None
    try:
        valid = train.dropna(subset=['홈팀득점', '원정팀득점'])
        if len(valid) >= 20:
            X_s = X_scaled.loc[valid.index]
            # [수정] 스코어 예측 모델도 경량화
            hg_model = RandomForestRegressor(n_estimators=50, max_depth=3, n_jobs=1, random_state=42).fit(X_s, valid['홈팀득점'])
            ag_model = RandomForestRegressor(n_estimators=50, max_depth=3, n_jobs=1, random_state=42).fit(X_s, valid['원정팀득점'])
    except: pass
    
    return model, le, hg_model, ag_model, features, scaler

def calc_handi_probs_and_ev(h_exp, a_exp, handi_val, h_odd, d_odd, l_odd):
    prob_h_win = 0; prob_draw = 0; prob_a_win = 0
    for h in range(10): 
        for a in range(10):
            p = poisson.pmf(h, h_exp) * poisson.pmf(a, a_exp)
            h_final = h + handi_val
            if h_final > a + 0.01: prob_h_win += p
            elif abs(h_final - a) < 0.01: prob_draw += p
            else: prob_a_win += p
    ev_win = (h_odd * prob_h_win) if pd.notnull(h_odd) else 0
    return ev_win, prob_h_win, prob_draw, prob_a_win

# ==========================================
# [3] 메인 웹앱 UI
# ==========================================
def main():
    st.title("⚽ AI 축구 분석 시스템 (Android Ver.)")
    
    # 1. 사이드바 - 설정
    st.sidebar.header("⚙️ 분석 설정")
    safe_prob = st.sidebar.slider("안전픽 최소 확률", 0.5, 0.99, 0.60, 0.01)
    min_ev = st.sidebar.number_input("가치픽 최소 EV", 0.5, 5.0, 1.05, 0.01)
    
    # 2. 파일 업로드
    uploaded_file = st.file_uploader("엑셀 파일(.xlsx)을 업로드하세요", type=['xlsx', 'xls'])
    
    # 세션 상태 초기화 (데이터 및 모델 유지)
    if 'data_loader' not in st.session_state: st.session_state.data_loader = DataLoader()
    if 'df' not in st.session_state: st.session_state.df = None
    if 'model_pack' not in st.session_state: st.session_state.model_pack = None

    if uploaded_file is not None:
        # 파일이 새로 올라오면 데이터 로드
        if st.session_state.df is None:
            with st.spinner("데이터 읽는 중..."):
                raw_df = st.session_state.data_loader.load_excel_data(uploaded_file)
                if raw_df is not None:
                    st.session_state.df = engineer_features(raw_df)
                    st.success(f"데이터 로드 완료! 총 {len(st.session_state.df)}경기")
        
        # 모델 학습 버튼
        if st.button("🚀 모델 학습 시작"):
            with st.spinner("AI가 학습 중입니다..."):
                res = train_models_func(st.session_state.df)
                if res:
                    st.session_state.model_pack = res
                    st.success("모델 학습 완료!")
                else:
                    st.error("학습 실패: 데이터가 부족하거나 오류가 있습니다.")

    # 분석 결과 출력
    if st.session_state.model_pack is not None:
        st.divider()
        st.subheader("📊 경기 분석 결과")
        
        model, le, hm, am, ft, sc = st.session_state.model_pack
        df = st.session_state.df
        sched = df[df['경기유형'] == '예정'].sort_values('datetime')
        
        if sched.empty:
            st.info("분석할 예정 경기가 없습니다.")
            return

        # 탭 메뉴 생성
        tab1, tab2, tab3 = st.tabs(["🏆 최종공통픽", "📋 전체 예측", "🛡️ 안전/가치픽"])
        
        with tab1:
            st.write("### AI 추천: 공통/교집합 픽")
            cnt = 0
            for idx, row in sched.iterrows():
                # 예측 로직
                X_df = pd.DataFrame([row[ft].fillna(0)])
                X_s = sc.transform(X_df)
                X_f = pd.DataFrame(X_s, columns=ft)
                
                probs = model.predict_proba(X_f)[0]
                pred = le.inverse_transform([model.predict(X_f)[0]])[0]
                hg = hm.predict(X_f)[0] if hm else 0
                ag = am.predict(X_f)[0] if am else 0
                
                # 변수 설정
                prob_map = {c:p for c,p in zip(le.classes_, probs)}
                p_win = prob_map.get('승', 0)
                ev_win = row['승'] * p_win
                
                is_fav = (pred == '승' and row['승'] < 1.75) or (pred == '패' and row['패'] < 1.75)
                is_safe = is_fav and (prob_map.get(pred, 0) >= safe_prob)
                is_high = (ev_win >= min_ev and p_win >= 0.30)
                
                trap_msg = st.session_state.data_loader.check_trap(row['Home'], row['Away'], row['승'])
                is_trap = True if trap_msg else False
                
                key = f"{row['Home']}vs{row['Away']}"
                c_info = st.session_state.data_loader.color_map.get(key)
                has_color = False
                if c_info:
                     if c_info.get('N_W') or c_info.get('N_L') or c_info.get('H_W') or c_info.get('H_L'):
                        has_color = True

                # 출력 조건
                msg = ""
                icon = ""
                if is_safe and is_high and not is_trap:
                    icon = "🌟"
                    msg = "1티어 (안전+가치)"
                elif is_safe and has_color and not is_trap:
                    icon = "✅"
                    msg = "2티어 (안전+컬러)"
                elif is_high and has_color and not is_trap:
                    icon = "💎"
                    msg = "3티어 (가치+컬러)"
                
                if msg:
                    cnt += 1
                    with st.expander(f"{icon} {row['Home']} vs {row['Away']} : {pred}"):
                        st.write(f"**유형:** {msg}")
                        st.write(f"**배당:** 승{row['승']} / 무{row['무']} / 패{row['패']}")
                        st.write(f"**예상 스코어:** {hg:.1f} : {ag:.1f}")
                        st.write(f"**근거:** 확률 {prob_map.get(pred)*100:.1f}% / EV {ev_win:.2f}")
                        if has_color: st.info("히든 컬러 감지됨")
            if cnt == 0:
                st.write("조건에 맞는 추천 경기가 없습니다.")

        with tab2:
            st.write("### 전체 경기 리스트")
            # 전체 데이터를 표로 보여주기 위한 처리
            all_results = []
            for idx, row in sched.iterrows():
                X_df = pd.DataFrame([row[ft].fillna(0)])
                X_s = sc.transform(X_df)
                X_f = pd.DataFrame(X_s, columns=ft)
                pred = le.inverse_transform([model.predict(X_f)[0]])[0]
                probs = model.predict_proba(X_f)[0]
                accuracy = max(probs)
                
                all_results.append({
                    '시간': row['경기시간'],
                    '홈팀': row['Home'],
                    '원정팀': row['Away'],
                    'AI픽': pred,
                    '확률': f"{accuracy*100:.1f}%",
                    '배당(승)': row['승']
                })
            st.dataframe(pd.DataFrame(all_results))

        with tab3:
            st.write("### 안전 / 가치 / 주의 경기")
            col1, col2 = st.columns(2)
            with col1:
                st.write("🛡️ **안전픽 (확률 높음)**")
                for idx, row in sched.iterrows():
                    X_df = pd.DataFrame([row[ft].fillna(0)])
                    X_s = sc.transform(X_df)
                    X_f = pd.DataFrame(X_s, columns=ft)
                    pred = le.inverse_transform([model.predict(X_f)[0]])[0]
                    probs = model.predict_proba(X_f)[0]
                    prob_val = max(probs)
                    
                    is_fav = (pred == '승' and row['승'] < 1.7) or (pred == '패' and row['패'] < 1.7)
                    if is_fav and prob_val >= safe_prob:
                        st.success(f"{row['Home']} vs {row['Away']} -> {pred} ({prob_val*100:.1f}%)")

            with col2:
                st.write("💎 **가치픽 (배당 대비 좋음)**")
                for idx, row in sched.iterrows():
                    X_df = pd.DataFrame([row[ft].fillna(0)])
                    X_s = sc.transform(X_df)
                    X_f = pd.DataFrame(X_s, columns=ft)
                    probs = model.predict_proba(X_f)[0]
                    prob_map = {c:p for c,p in zip(le.classes_, probs)}
                    p_win = prob_map.get('승', 0)
                    ev_win = row['승'] * p_win
                    
                    if ev_win >= min_ev and p_win >= 0.30:
                        st.warning(f"{row['Home']} vs {row['Away']} -> 승 (EV {ev_win:.2f})")

if __name__ == '__main__':

    main()
